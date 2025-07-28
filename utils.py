import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def load_config(file_path):
    """Load configuration file into DataFrame"""
    return pd.read_excel(file_path)

# =================== TEAM ===================

def get_team_list(team_df):
    """Lấy danh sách nhóm trưởng"""
    return team_df['Group Leader'].dropna().unique().tolist()

def get_team_members(team_df, group_leader):
    """Lấy danh sách thành viên theo nhóm trưởng"""
    members = team_df[team_df['Group Leader'] == group_leader]['Employee Name']
    return members.dropna().unique().tolist()

def get_employee_id(team_df, employee_name):
    """Trả về mã nhân viên từ tên"""
    match = team_df[team_df['Employee Name'] == employee_name]
    if not match.empty:
        return match.iloc[0]['Employee ID']
    return "N/A"

# =================== PROJECT ===================

def get_project_list(project_df):
    """Lấy danh sách tên dự án"""
    return project_df['Project Name'].dropna().unique().tolist()

def get_project_code(project_df, project_name):
    """Trả về mã dự án từ tên"""
    match = project_df[project_df['Project Name'] == project_name]
    if not match.empty:
        return match.iloc[0]['Project Code']
    return "N/A"

# =================== JOB ===================

def get_job_list(job_df, selected_project_code_or_name):
    """Lọc danh sách công việc theo mã dự án hoặc tên dự án"""
    if 'Project Code' in job_df.columns:
        mask = (job_df['Project Code'] == selected_project_code_or_name) | \
               (job_df['Project Name'] == selected_project_code_or_name)
    else:
        mask = job_df['Project Name'] == selected_project_code_or_name
    return job_df[mask]['Job Name'].dropna().unique().tolist()

def get_job_info(job_df, job_name):
    """Trả về (Job Code, Workcenter, Task) từ Job Name"""
    match = job_df[job_df['Job Name'] == job_name]
    if not match.empty:
        row = match.iloc[0]
        return (
            row.get('Job Code', 'N/A'),
            row.get('Workcenter', 'N/A'),
            row.get('Task', 'N/A')
        )
    return ("N/A", "N/A", "N/A")

def save_to_time_report(data_list, file_path='Time_report.xlsm', sheet_name='Raw Data'):
    """Ghi danh sách dữ liệu vào sheet Raw Data trong Time_report.xlsm."""

    columns = [
        "Date", "Project Name", "Project Code", "Job Name", "Job Code",
        "Employee", "Employee ID", "Team", "Workcenter", "Task", "Hours"
    ]

    # Kiểm tra file đã tồn tại chưa
    file_exists = os.path.exists(file_path)

    # Nếu chưa có thì tạo mới và thêm header
    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(columns)
        wb.save(file_path)

    # Mở file (keep_vba=True nếu có macro)
    try:
        wb = load_workbook(file_path, keep_vba=True)
    except:
        wb = load_workbook(file_path)

    # Lấy hoặc tạo sheet
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)
    else:
        ws = wb[sheet_name]

    # Tìm dòng đầu tiên trống
    next_row = ws.max_row + 1
    if all([cell.value is None for cell in ws[next_row - 1]]):
        next_row -= 1

    # Ghi từng dòng dữ liệu
    for entry in data_list:
        row_data = [entry.get(col, "") for col in columns]
        ws.append(row_data)

    wb.save(file_path)
    return file_path
